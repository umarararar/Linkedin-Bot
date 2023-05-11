from selenium import webdriver
import openpyxl
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait

#PATH = "chromedriver.exe"

options = Options()
options.add_argument("--user-data-dir=C:/Users/Umar/AppData/Local/Google/Chrome/User Data")
options.add_argument("profile-directory=Profile 6")

driver = webdriver.Chrome(options=options)
#driver = uc.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(10)  # wait for 10 seconds
driver.get("https://www.linkedin.com")


workbook = openpyxl.load_workbook('Book2.xlsx')
worksheet = workbook['Sheet1']
# link = worksheet.cell(row=1, column=1).value
# print(link)
for row in worksheet.iter_rows(min_row=5, values_only=True):
    # Get the value of the first cell (assuming it's a link)
    link = row[0]
    
    # Navigate to the link
    driver.get(link)
    wait = WebDriverWait(driver, 10)
    sleep(4)
   #ember88 > span
   #/html/body/div[5]/div[3]/div/div/div[2]/div/div/main/section[1]/div[2]/div[3]/div/button/span
   #ember128 > span
   #/html/body/div[5]/div[3]/div/div/div[2]/div/div/main/section[1]/div[2]/div[3]/div/button/span
   #//*[@id="ember128"]/span
   #//*[@id="ember99"]/span
   #//*[@id="ember97"]/span
   
   #//*[@id="ember80"]/span

# import re

# # Define the pattern to match the Connect button's XPath
# pattern = r'^#\/\*[@id="ember\d+"]\/span$'

# # Iterate through all elements on the page
# for conn in driver.find_elements(By.XPATH,'//*'):
#     # Check if the element's XPath matches the pattern
#     if re.match(pattern, conn.get_attribute('id')):
#         # Click the Connect button
#         if "Connect" in conn.text:
#           conn.click()
#           sleep(1)
#           try:
#             send_button = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div[3]/button[2]/span")))
#             send_button.click()
#             sleep(2)

#           except:
#             pass
#         else:
#           continue
          
#          # break   
    try:
      connect_button1 = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[5]/div[3]/div/div/div[2]/div/div/main/section[1]/div[2]/div[3]/div/button/span")))

      if "Connect" in connect_button1.text:

        driver.execute_script("arguments[0].click();", connect_button1)

        sleep(1)
        try:
          send_button = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[3]/div/div/div[3]/button[2]/span")))
          send_button.click()
          sleep(2)

        except:
          pass
      else:
        continue
    except:
      pass

